"""Renders a DayReport (see backup_report.py) into an Excel workbook that
visually matches the "Export Excel" button on the Expenses page — same
branch item blocks (opening/delivery/price/closing/used/expense) side by
side, followed by a sales/bills/stock-expense/profit summary table. See
writeBranchBlock/buildDaySheet in AdminExpensesView.vue for the original.

Built entirely in memory (BytesIO) — never touches disk.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.backup_report import DayReport

EXPORT_COLS = 7
BRANCHES_PER_ROW = 2
GAP_COLS = 1
BLOCK_WIDTH = EXPORT_COLS + GAP_COLS

ORANGE = "FFFFA733"
YELLOW = "FFFFF200"
RED = "FFFF4D4D"
GREEN = "FF1B7A43"
DARK_RED = "FFCC0000"
BORDER_GRAY = "FF999999"

_thin = Side(style="thin", color=BORDER_GRAY)
ALL_BORDERS = Border(top=_thin, bottom=_thin, left=_thin, right=_thin)

TITLE_FILL = PatternFill("solid", fgColor=ORANGE)
HEADER_FILL = PatternFill("solid", fgColor=YELLOW)
TOTAL_FILL = PatternFill("solid", fgColor=YELLOW)

TITLE_FONT = Font(bold=True, size=13)
HEADER_FONT = Font(bold=True, color=DARK_RED)
TOTAL_FONT = Font(bold=True)

ALIGN_LEFT = Alignment(vertical="center", horizontal="left")
ALIGN_RIGHT = Alignment(vertical="center", horizontal="right")
ALIGN_CENTER = Alignment(vertical="center", horizontal="center")


def peso(amount: float | None) -> str:
    return f"₱{(amount or 0):,.2f}"


def _set_cell(sheet: Worksheet, row: int, col: int, value, *, fill=None, font=None, align=None, border=ALL_BORDERS):
    cell = sheet.cell(row=row + 1, column=col + 1, value=value)
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border
    return cell


def _write_branch_block(sheet: Worksheet, group, row_offset: int, col_offset: int) -> int:
    r = row_offset

    _set_cell(sheet, r, col_offset, group.name, fill=TITLE_FILL, font=TITLE_FONT, align=ALIGN_LEFT)
    for c in range(1, EXPORT_COLS - 1):
        _set_cell(sheet, r, col_offset + c, "", fill=TITLE_FILL, font=TITLE_FONT, align=ALIGN_LEFT)
    _set_cell(sheet, r, col_offset + EXPORT_COLS - 1, peso(group.total), fill=TITLE_FILL, font=TITLE_FONT, align=ALIGN_RIGHT)
    sheet.merge_cells(
        start_row=r + 1, start_column=col_offset + 1, end_row=r + 1, end_column=col_offset + EXPORT_COLS - 1
    )
    r += 1

    for i, header in enumerate(["Item", "Opening", "Delivery", "Price", "Closing", "Used", "Expense"]):
        _set_cell(sheet, r, col_offset + i, header, fill=HEADER_FILL, font=HEADER_FONT, align=ALIGN_CENTER)
    r += 1

    for row in group.rows:
        label = f"{row.item_name} {row.unit}" if row.unit else row.item_name
        _set_cell(sheet, r, col_offset, label, align=ALIGN_LEFT)
        _set_cell(sheet, r, col_offset + 1, row.opening, align=ALIGN_RIGHT)
        _set_cell(sheet, r, col_offset + 2, row.delivery, align=ALIGN_RIGHT)
        _set_cell(sheet, r, col_offset + 3, peso(row.price), align=ALIGN_RIGHT)
        _set_cell(sheet, r, col_offset + 4, row.closing if row.has_closing else "—", align=ALIGN_RIGHT)

        used_font = None
        if row.has_closing and row.used is not None:
            if row.used < 0:
                used_font = Font(color=RED)
            elif row.used > 0:
                used_font = Font(color=GREEN)
        _set_cell(sheet, r, col_offset + 5, row.used if row.has_closing else "—", font=used_font, align=ALIGN_RIGHT)

        expense_font = Font(color=RED) if row.has_closing and (row.expense or 0) < 0 else None
        _set_cell(
            sheet, r, col_offset + 6, peso(row.expense) if row.has_closing else "—", font=expense_font, align=ALIGN_RIGHT
        )
        r += 1

    _set_cell(sheet, r, col_offset, "TOTAL EXPENSES", fill=TOTAL_FILL, font=TOTAL_FONT, align=ALIGN_LEFT)
    for c in range(1, EXPORT_COLS - 1):
        _set_cell(sheet, r, col_offset + c, "", fill=TOTAL_FILL, font=TOTAL_FONT, align=ALIGN_LEFT)
    _set_cell(sheet, r, col_offset + EXPORT_COLS - 1, peso(group.total), fill=TOTAL_FILL, font=TOTAL_FONT, align=ALIGN_RIGHT)
    sheet.merge_cells(
        start_row=r + 1, start_column=col_offset + 1, end_row=r + 1, end_column=col_offset + EXPORT_COLS - 1
    )
    r += 1

    return r - row_offset


def _build_day_sheet(sheet: Worksheet, report: DayReport) -> None:
    row_offset = 0
    max_col = 0

    groups = report.groups
    for i in range(0, len(groups), BRANCHES_PER_ROW):
        chunk = groups[i : i + BRANCHES_PER_ROW]
        chunk_height = 0
        for idx, group in enumerate(chunk):
            col_offset = idx * BLOCK_WIDTH
            height = _write_branch_block(sheet, group, row_offset, col_offset)
            chunk_height = max(chunk_height, height)
            max_col = max(max_col, col_offset + EXPORT_COLS - 1)
        row_offset += chunk_height + 1

    summary_start_row = row_offset
    _set_cell(sheet, row_offset, 0, f"Daily Sales & Bills — {report.date.isoformat()}", fill=TITLE_FILL, font=TITLE_FONT, align=ALIGN_LEFT)
    for c in range(1, EXPORT_COLS):
        _set_cell(sheet, row_offset, c, "", fill=TITLE_FILL, font=TITLE_FONT, align=ALIGN_LEFT)
    sheet.merge_cells(
        start_row=summary_start_row + 1, start_column=1, end_row=summary_start_row + 1, end_column=EXPORT_COLS
    )
    row_offset += 1

    for i, header in enumerate(["Branch", "Sales", "Daily Bills", "Stock Expense", "Total Expense", "Profit", ""]):
        _set_cell(sheet, row_offset, i, header, fill=HEADER_FILL, font=HEADER_FONT, align=ALIGN_CENTER)
    row_offset += 1

    for row in report.summaries:
        profit_font = Font(color=GREEN if row.profit >= 0 else RED)
        _set_cell(sheet, row_offset, 0, row.branch_name, align=ALIGN_LEFT)
        _set_cell(sheet, row_offset, 1, peso(row.sales), align=ALIGN_RIGHT)
        _set_cell(sheet, row_offset, 2, peso(row.bills), align=ALIGN_RIGHT)
        _set_cell(sheet, row_offset, 3, peso(row.stock_expense), align=ALIGN_RIGHT)
        _set_cell(sheet, row_offset, 4, peso(row.total_expense), align=ALIGN_RIGHT)
        _set_cell(sheet, row_offset, 5, peso(row.profit), font=profit_font, align=ALIGN_RIGHT)
        _set_cell(sheet, row_offset, 6, "", align=ALIGN_LEFT)
        row_offset += 1

    totals = report.totals
    totals_profit_font = Font(bold=True, color=GREEN if totals["profit"] >= 0 else RED)
    _set_cell(sheet, row_offset, 0, "TOTAL", fill=TOTAL_FILL, font=TOTAL_FONT, align=ALIGN_LEFT)
    _set_cell(sheet, row_offset, 1, peso(totals["sales"]), fill=TOTAL_FILL, font=TOTAL_FONT, align=ALIGN_RIGHT)
    _set_cell(sheet, row_offset, 2, peso(totals["bills"]), fill=TOTAL_FILL, font=TOTAL_FONT, align=ALIGN_RIGHT)
    _set_cell(sheet, row_offset, 3, peso(totals["stock_expense"]), fill=TOTAL_FILL, font=TOTAL_FONT, align=ALIGN_RIGHT)
    _set_cell(sheet, row_offset, 4, peso(totals["total_expense"]), fill=TOTAL_FILL, font=TOTAL_FONT, align=ALIGN_RIGHT)
    _set_cell(sheet, row_offset, 5, peso(totals["profit"]), fill=TOTAL_FILL, font=totals_profit_font, align=ALIGN_RIGHT)
    _set_cell(sheet, row_offset, 6, "", fill=TOTAL_FILL, font=TOTAL_FONT, align=ALIGN_LEFT)

    num_group_cols = min(len(groups), BRANCHES_PER_ROW) or 1
    widths = [18, 10, 10, 12, 10, 10, 12]
    col = 1
    for g in range(num_group_cols):
        for w in widths:
            sheet.column_dimensions[get_column_letter(col)].width = w
            col += 1
        if g < num_group_cols - 1:
            sheet.column_dimensions[get_column_letter(col)].width = 3
            col += 1


def _day_tab_name(report: DayReport) -> str:
    return f"{report.date.strftime('%b')} {report.date.day}"


def _build_summary_sheet(sheet: Worksheet, reports: list[DayReport], month_label: str) -> None:
    cols = 6
    r = 0

    _set_cell(sheet, r, 0, f"{month_label} — Daily Summary (click a date to jump to its sheet)", fill=TITLE_FILL, font=TITLE_FONT, align=ALIGN_LEFT)
    for c in range(1, cols):
        _set_cell(sheet, r, c, "", fill=TITLE_FILL, font=TITLE_FONT, align=ALIGN_LEFT)
    sheet.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=cols)
    r += 1

    for i, header in enumerate(["Date", "Sales", "Daily Bills", "Stock Expense", "Total Expense", "Profit"]):
        _set_cell(sheet, r, i, header, fill=HEADER_FILL, font=HEADER_FONT, align=ALIGN_CENTER)
    r += 1

    month_totals = {"sales": 0.0, "bills": 0.0, "stock_expense": 0.0, "total_expense": 0.0, "profit": 0.0}
    for report in reports:
        tab_name = _day_tab_name(report)
        totals = report.totals
        link_cell = _set_cell(sheet, r, 0, tab_name, align=ALIGN_LEFT)
        link_cell.hyperlink = f"#'{tab_name}'!A1"
        link_cell.font = Font(color="FF1155CC", underline="single", bold=True)
        _set_cell(sheet, r, 1, peso(totals["sales"]), align=ALIGN_RIGHT)
        _set_cell(sheet, r, 2, peso(totals["bills"]), align=ALIGN_RIGHT)
        _set_cell(sheet, r, 3, peso(totals["stock_expense"]), align=ALIGN_RIGHT)
        _set_cell(sheet, r, 4, peso(totals["total_expense"]), align=ALIGN_RIGHT)
        profit_font = Font(color=GREEN if totals["profit"] >= 0 else RED)
        _set_cell(sheet, r, 5, peso(totals["profit"]), font=profit_font, align=ALIGN_RIGHT)
        for key in month_totals:
            month_totals[key] = round(month_totals[key] + totals[key] + 1e-9, 2)
        r += 1

    totals_profit_font = Font(bold=True, color=GREEN if month_totals["profit"] >= 0 else RED)
    _set_cell(sheet, r, 0, "MONTH TOTAL", fill=TOTAL_FILL, font=TOTAL_FONT, align=ALIGN_LEFT)
    _set_cell(sheet, r, 1, peso(month_totals["sales"]), fill=TOTAL_FILL, font=TOTAL_FONT, align=ALIGN_RIGHT)
    _set_cell(sheet, r, 2, peso(month_totals["bills"]), fill=TOTAL_FILL, font=TOTAL_FONT, align=ALIGN_RIGHT)
    _set_cell(sheet, r, 3, peso(month_totals["stock_expense"]), fill=TOTAL_FILL, font=TOTAL_FONT, align=ALIGN_RIGHT)
    _set_cell(sheet, r, 4, peso(month_totals["total_expense"]), fill=TOTAL_FILL, font=TOTAL_FONT, align=ALIGN_RIGHT)
    _set_cell(sheet, r, 5, peso(month_totals["profit"]), fill=TOTAL_FILL, font=totals_profit_font, align=ALIGN_RIGHT)

    widths = [14, 12, 12, 14, 14, 12]
    for i, w in enumerate(widths):
        sheet.column_dimensions[get_column_letter(i + 1)].width = w


def build_month_workbook(reports: list[DayReport], month_label: str) -> BytesIO:
    """Rebuilds the whole month's backup workbook from scratch every time it's
    called — one sheet per day (1st of the month through the latest day given)
    plus a Summary sheet — rather than trying to merge into a previously
    uploaded file. Simpler and immune to drift: the file always reflects
    exactly what's in the database right now, however many times a day it
    gets regenerated.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)

    for report in reports:
        sheet = workbook.create_sheet(title=_day_tab_name(report))
        _build_day_sheet(sheet, report)

    summary_sheet = workbook.create_sheet(title="Summary", index=0)
    _build_summary_sheet(summary_sheet, reports, month_label)
    workbook.active = 0

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
